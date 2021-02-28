import openpyxl
wk = openpyxl.load_workbook("D:\\Book1.xlsx")
sh = wk['Sheet1']
print(sh['A3'].value)  # fetch data from workBook
print(sh['B2'].value)

# Fetch data using cell numbers
cl = sh.cell(2, 2)
print(cl.value)

# Another Approch is

cl = sh.cell(row=3, column=2)  #
cl = sh.cell(column=1, row=3)

print(cl.row)
print(cl.column)
print(cl.value)

# Find complete Data from work book
rd = openpyxl.load_workbook("D:\\Book1.xlsx")
fd = rd['Sheet1']

rows = fd.max_row
column = fd.max_column

print("Total rows : -", str(rows))
print("Total column : -", str(column))

for r in sh['A1': 'D3']:  # Approch 2
    for c in r:
        print(c.value)


# for i in range(1, rows+1):
#     for j in range(1, column+1):
#         c = sh.cell(i, j)
#         # print(c)
#         print(c.value)
