import openpyxl

wk = openpyxl.Workbook()
sh = wk.active
sh.title = "New Sheet"
print(sh.title)

sh['A4'].value = 'Shoaib@gmail.com'
wk.create_sheet(title='Testing')  # tile="Shoaib WK"
# wk.save("D:\\writeDta.xlsx")

# second Sheet
wk.create_sheet(title='Testing')
sh1 = wk['Testing']
sh1['A3'] = '80997888'  # sh1['A3]="data entry to sheets"

# REMOVE SHEETS
wk.remove(wk['Testing'])

wk.save("D:\\writeDta.xlsx")
